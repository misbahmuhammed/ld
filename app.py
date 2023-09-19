from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
import pandas as pd
import os
import zipfile
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = 'misbah'
# Define a function to read data from the master data Excel
def read_master_data():
    try:
        if os.path.exists('master_data.xlsx'):
            df = pd.read_excel('master_data.xlsx', sheet_name='Master Data')
            return df.to_dict(orient='records')
        else:
            return []  # Return an empty list if the file doesn't exist
    except Exception as e:
        print(f"Error reading master data Excel: {str(e)}")
        return []

# Define a function to read data from the allocated courses Excel
def read_allocated_courses_data():
    try:
        if os.path.exists('allocated_courses.xlsx'):
            df = pd.read_excel('allocated_courses.xlsx', sheet_name='Allocated Courses')
            return df.to_dict(orient='records')
        else:
            return []  # Return an empty list if the file doesn't exist
    except Exception as e:
        print(f"Error reading allocated courses Excel: {str(e)}")
        return []

# ... (rest of your code remains the same)

@app.route('/allocate_course', methods=['GET', 'POST'])
def allocate_course():
    global selected_batch  # Declare selected_batch as global
    
    if request.method == 'POST':
        # Retrieve form data
        batch = request.form.get('batch')
        course_type = request.form.get('course_type')
        course_name = request.form.get('course_name')
        duration_start = request.form.get('duration_start')
        duration_end = request.form.get('duration_end')
        online_offline = request.form.get('online_offline')
        trainer = request.form.get('trainer')
        
        # Define column names
        columns = [
            'Batch',
            'Course Type',
            'Course Name',
            'Course Duration-Start',
            'Course Duration-End',
            'Online/Offline',
            'Trainer'
        ]

        # Create a DataFrame with your data and column names
        data = {
            'Batch': [batch],
            'Course Type': [course_type],
            'Course Name': [course_name],
            'Course Duration-Start': [duration_start],
            'Course Duration-End': [duration_end],
            'Online/Offline': [online_offline],
            'Trainer': [trainer]
        }
        df = pd.DataFrame(data, columns=columns)

        # Append to 'allocated_courses.xlsx'
        excel_file = 'allocated_courses.xlsx'
        if os.path.exists(excel_file):
            # Load the existing Excel file
            book = load_workbook(excel_file)
            writer = pd.ExcelWriter(excel_file, engine='openpyxl') 
            writer.book = book
        else:
            # Create a new Excel file if it doesn't exist
            writer = pd.ExcelWriter(excel_file, engine='openpyxl')
            # Create a new workbook with a sheet named 'Allocated Courses'
            writer.book = Workbook()
            writer.book.active.title = 'Allocated Courses'
        
        # Check if the 'Allocated Courses' sheet already exists
        if 'Allocated Courses' in writer.book.sheetnames:
            # Append the data to the existing sheet without the header
            startrow = writer.sheets['Allocated Courses'].max_row
            df.to_excel(writer, sheet_name='Allocated Courses', index=False, header=False, startrow=startrow)
        else:
            # Create the sheet and write the data with the header
            df.to_excel(writer, sheet_name='Allocated Courses', index=False, header=True)

        writer.save()
        writer.close()

    # Fetch Batch names and Course Types from master Excel and pass them to the template
    master_data = read_master_data()
    batch_names = [record['Batch'] for record in master_data]
    course_types = list(set(record['Course Type'] for record in master_data))

    return render_template('allocate_course.html', batch_names=batch_names, course_types=course_types, master_data=master_data)


@app.route('/dashboard')
def dashboard():
    # You can add other dashboard content here if needed
    return render_template('dashboard.html')



@app.route('/allocated_courses')
def allocated_courses():
    # Fetch data from the "allocated courses" Excel and pass it to the template
    allocated_courses_data = read_allocated_courses_data()
    return render_template('allocated_course.html', allocated_courses_data=allocated_courses_data)



@app.route('/get_course_and_trainer_options', methods=['GET'])
def get_course_and_trainer_options():
    selected_course_type = request.args.get('course_type')

    # Filter master_data to get course names and trainers based on the selected course type
    course_names = [record['Course Name'] for record in master_data if record['Course Type'] == selected_course_type]
    trainers = [record['Trainer'] for record in master_data if record['Course Type'] == selected_course_type]

    # Return the filtered course names and trainers as JSON response
    return jsonify(course_names=course_names, trainers=trainers)



@app.route('/add_new_course', methods=['GET', 'POST'])
def add_new_course():
    if request.method == 'POST':
        # Retrieve form data
        new_course_type = request.form.get('new_course_type')
        new_course_name = request.form.get('new_course_name')
        new_duration = request.form.get('new_duration')
        new_online_offline = request.form.get('new_online_offline')
        new_trainer_id = request.form.get('new_trainer_id')  # Retrieve Trainer ID

        # Create a DataFrame with the new course data
        new_course_data = {
            'Course Type': [new_course_type],
            'Course Name': [new_course_name],
            'Duration': [new_duration],
            'Online/Offline': [new_online_offline],
            'Trainer ID': [new_trainer_id]  # Store Trainer ID
        }

        # Load the existing 'master_data.xlsx' file
        book = load_workbook('master_data.xlsx')
        writer = pd.ExcelWriter('master_data.xlsx', engine='openpyxl') 
        writer.book = book

        # Check if the 'Master Data' sheet exists and if it's empty, add column names
        if 'Master Data' in writer.book.sheetnames and writer.sheets['Master Data'].max_row == 0:
            df_columns = [
                'Batch',
                'Course Type',
                'Course Name',
                'Course Duration-Start',
                'Course Duration-End',
                'Online/Offline',
                'Trainer',
                'Trainer ID',
            ]
            df = pd.DataFrame(columns=df_columns)
            df.to_excel(writer, sheet_name='Master Data', index=False)

        # Write the new course data to the 'Master Data' sheet
        df = pd.DataFrame(new_course_data)
        df.to_excel(writer, sheet_name='Master Data', index=False, header=False, startrow=writer.sheets['Master Data'].max_row + 1)

        # Save and close the writer
        writer.save()
        writer.close()

        # Add a success message to be displayed on the dashboard
        success_message = "Course added successfully!"

        # Fetch data from the master data Excel
        master_data = read_master_data()

        return render_template('dashboard.html', master_data=master_data, success_message=success_message)

    return render_template('add_new_course.html')

@app.route('/get_course_names', methods=['GET'])
def get_course_names():
    selected_course_type = request.args.get('course_type')

    # Filter course names based on the selected course type
    course_names = [record['Course Name'] for record in master_data if record['Course Type'] == selected_course_type]

    return jsonify(course_names=course_names)


# Flask route to fetch unique course types
@app.route('/get_course_types')
def get_course_types():
    master_data = read_master_data()
    course_types = list(set(record['Course Type'] for record in master_data))
    return jsonify(course_types=course_types)

@app.route('/get_trainers', methods=['GET'])
def get_trainers():
    selected_course_name = request.args.get('course_name')

    # Fetch trainers based on the selected course name (implement this logic)
    # trainers = ...

    # For example, assuming trainers is a list of trainer names
    trainers = ["Trainer 1", "Trainer 2", "Trainer 3"]

    return jsonify(trainers=trainers)

# ...



if __name__ == '__main__':

    app.run(debug=True)    